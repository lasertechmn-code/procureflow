import openpyxl, json, re
from collections import OrderedDict

PATH = r'historicData\Submitted PRs\MasterList\Master_PR_Record.xlsm'
wb = openpyxl.load_workbook(PATH, data_only=True, keep_vba=False)
ws = wb['TEMPLATE']

HEADERS = ['Requestor','Item','Details','Type','Buildcell','Company','Model','Hyperlink','Cost','Unit','Qty','SubTotal','Ordered','Received']

rows = []
for r in range(4, ws.max_row+1):
    cells = [ws.cell(row=r, column=c) for c in range(1,15)]
    vals = [c.value for c in cells]
    if all(v is None or (isinstance(v,str) and v.strip()=='') for v in vals):
        continue
    hl_cell = cells[7]
    link_target = hl_cell.hyperlink.target if hl_cell.hyperlink else None
    item_link = cells[1].hyperlink.target if cells[1].hyperlink else None
    rec = OrderedDict(zip(HEADERS, vals))
    rec['LinkTarget'] = link_target
    rec['ItemLink'] = item_link
    rec['_row'] = r
    rows.append(rec)

ALIASES = {
    'svenalagic':('Sven','Alagic'), 'sven':('Sven','Alagic'),
    'darrinloehr':('Darrin','Loehr'), 'darrin':('Darrin','Loehr'), 'loehr':('Darrin','Loehr'),
    'gerald':('Gerald','Jackson'), 'geraldjackson':('Gerald','Jackson'),
    'maureenkapler':('Maureen','Kapler'),
    'morganelliott':('Morgan','Elliott'), 'morgan':('Morgan','Elliott'),
    'tomdemars':('Tom','DeMars'), 'tomd':('Tom','DeMars'),
    'danpast':('Dan','Past'),
    'mikezavertaniy':('Mike','Zavertaniy'),
    'romandavtyan':('Roman','Davtyan'), 'roman':('Roman','Davtyan'), 'romandavtyna':('Roman','Davtyan'),
    'sydneythompson':('Sydney','Thompson'), 'sydney':('Sydney','Thompson'),
    'chufu':('Chufu',''),
    'jeremywinter':('Jeremy','Winter'),
    'mikeg':('Mike','Greere'),
    'toddleadens':('Todd','Leadens'),
    'macleod':('','Macleod'),
    'hanson':('','Hanson'),
    'gregzdeb':('Greg','Zdeb'), 'zdeb':('Greg','Zdeb'),
}

def parse_requestor(s):
    s = str(s).strip()
    date_str=None; time_str=None
    m = re.search(r'(\d{8})(?:_(\d{6}))?\s*$', s)
    namepart = s
    if m:
        d = m.group(1)
        mm,dd,yyyy = d[0:2],d[2:4],d[4:8]
        date_str=f"{yyyy}-{mm}-{dd}"
        if m.group(2):
            t=m.group(2); time_str=f"{t[0:2]}:{t[2:4]}:{t[4:6]}"
        namepart = s[:m.start()].rstrip('_ ')
    namepart = re.sub(r'^[Cc]\.?\s*', '', namepart)
    namepart = re.sub(r'\(\d+\)', '', namepart)
    key = re.sub(r'[^a-z]', '', namepart.lower())
    fn,ln = ALIASES.get(key, (None,None))
    if fn is None:
        parts = re.split(r'[ _]+', namepart.strip())
        if len(parts)>=2: fn,ln = parts[0], ' '.join(parts[1:])
        else: fn,ln = (parts[0] if parts else namepart), ''
    return fn, ln, date_str, time_str

for rec in rows:
    fn,ln,d,t = parse_requestor(rec['Requestor'])
    rec['ReqFirst']=fn; rec['ReqLast']=ln; rec['ReqDate']=d; rec['ReqTime']=t

def keyof(s):
    s=re.sub(r'(\d{8})(?:_(\d{6}))?\s*$','',str(s)).rstrip('_ ')
    s=re.sub(r'^[Cc]\.?\s*','',s); s=re.sub(r'\(\d+\)','',s)
    return re.sub(r'[^a-z]','',s.lower())
unmapped = sorted({rec['Requestor'] for rec in rows if keyof(rec['Requestor']) not in ALIASES})
print("RAW NOT IN ALIAS MAP (fallback used):")
for u in unmapped: print("   ", repr(u))

people = OrderedDict()
for rec in rows:
    name=(rec['ReqFirst'],rec['ReqLast'])
    people[name]=people.get(name,0)+1
print("\nCANONICAL PEOPLE:", len(people))
for (fn,ln),c in sorted(people.items()):
    print(f"  [{c:4d}] {fn!r} {ln!r}")

companies = OrderedDict()
for rec in rows:
    co = rec['Company']
    if co: companies[str(co).strip()] = companies.get(str(co).strip(),0)+1
print("\nUNIQUE COMPANIES:", len(companies))
for k,v in sorted(companies.items()): print(f"  [{v:4d}] {k!r}")

print("\nTYPES:", sorted({str(r['Type']).strip() for r in rows if r['Type']}))
print("\nBUILDCELLS:", sorted({str(r['Buildcell']).strip() for r in rows if r['Buildcell']}))
print("\nUNITS:", sorted({str(r['Unit']).strip() for r in rows if r['Unit']}))
print("\nORDERED VALUES:", sorted({str(r['Ordered']).strip() for r in rows if r['Ordered'] is not None}))
print("\nRECEIVED VALUES:", sorted({str(r['Received']).strip() for r in rows if r['Received'] is not None}))

groups = OrderedDict()
for rec in rows:
    k=str(rec['Requestor'])
    groups.setdefault(k,[]).append(rec)
print("\nPR GROUPS (by raw requestor):", len(groups))

with open(r'scripts\extracted.json','w',encoding='utf-8') as f:
    json.dump(rows, f, default=str, indent=1)
print("\nWrote scripts/extracted.json with", len(rows), "rows")
