import openpyxl, json, re
from collections import OrderedDict, Counter
from datetime import datetime, timedelta

PATH = r'historicData\Submitted PRs\MasterList\Master_PR_Record.xlsm'
wb = openpyxl.load_workbook(PATH, data_only=True, keep_vba=False)
ws = wb['TEMPLATE']
HEADERS = ['Requestor','Item','Details','Type','Buildcell','Company','Model','Hyperlink','Cost','Unit','Qty','SubTotal','Ordered','Received']

rows = []
for r in range(4, ws.max_row+1):
    cells = [ws.cell(row=r, column=c) for c in range(1,15)]
    vals = [c.value for c in cells]
    if all(v is None or (isinstance(v,str) and v.strip()=='') for v in vals):
        continue
    rec = OrderedDict(zip(HEADERS, vals))
    rec['LinkTarget'] = cells[7].hyperlink.target if cells[7].hyperlink else None
    rec['ItemLink'] = cells[1].hyperlink.target if cells[1].hyperlink else None
    rows.append(rec)

# ---------- Requestor normalization ----------
ALIASES = {
    'svenalagic':('Sven','Alagic'), 'sven':('Sven','Alagic'),
    'darrinloehr':('Darrin','Loehr'), 'darrin':('Darrin','Loehr'), 'loehr':('Darrin','Loehr'),
    'gerald':('Gerald','Jackson'), 'geraldjackson':('Gerald','Jackson'),
    'maureenkapler':('Maureen','Kapler'),
    'morganelliott':('Morgan','Elliott'), 'morgan':('Morgan','Elliott'),
    'tomdemars':('Tom','DeMars'), 'tomd':('Tom','DeMars'),
    'danpast':('Dan','Past'),
    'mikezavertaniy':('Mike','Zavertaniy'),
    'romandavtyan':('Roman','Davtyan'), 'roman':('Roman','Davtyan'), 'romandavtyna':('Roman','Davtyan'),
    'sydneythompson':('Sydney','Thompson'), 'sydney':('Sydney','Thompson'),
    'chufu':('Chufu','Lin'),
    'jeremywinter':('Jeremy','Winter'),
    'mikeg':('Mike','Greere'),
    'toddleadens':('Todd','Leadens'),
    'macleod':('Macleod','Macleod'),
    'hanson':('Hanson','Hanson'),
    'gregzdeb':('Greg','Zdeb'), 'zdeb':('Greg','Zdeb'),
}

def parse_requestor(s):
    s = str(s).strip()
    date_str=None; time_str=None
    m = re.search(r'(\d{8})(?:_(\d{6}))?\s*$', s)
    namepart = s
    if m:
        d = m.group(1)
        mm,dd,yyyy = d[0:2],d[2:4],d[4:8]
        date_str=f"{yyyy}-{mm}-{dd}"
        if m.group(2):
            t=m.group(2); time_str=f"{t[0:2]}:{t[2:4]}:{t[4:6]}"
        namepart = s[:m.start()].rstrip('_ ')
    namepart = re.sub(r'^[Cc]\.\s*', '', namepart)   # only strip "C." prefix, keep "Chufu"
    namepart = re.sub(r'\(\d+\)', '', namepart)
    key = re.sub(r'[^a-z]', '', namepart.lower())
    fn,ln = ALIASES.get(key, (None,None))
    if fn is None:
        parts = re.split(r'[ _]+', namepart.strip())
        fn,ln = (parts[0], ' '.join(parts[1:])) if len(parts)>=2 else (parts[0] if parts else namepart, '')
    return fn, ln, date_str, time_str

# ---------- Company normalization + landing pages ----------
VENDOR_URLS = {
    'amazon': ('Amazon','https://www.amazon.com'),
    'automationdirect': ('Automation Direct','https://www.automationdirect.com'),
    'digikey': ('DigiKey','https://www.digikey.com'),
    'fisherscientific': ('Fisher Scientific','https://www.fishersci.com'),
    'globalindustrial': ('Global Industrial','https://www.globalindustrial.com'),
    'grainger': ('Grainger','https://www.grainger.com'),
    'kurtjlesker': ('Kurt J. Lesker','https://www.lesker.com'),
    'lakeshore': ('Lake Shore','https://www.lakeshore.com'),
    'mscdirect': ('MSC Direct','https://www.mscdirect.com'),
    'msc': ('MSC Direct','https://www.mscdirect.com'),
    'mcmastercarr': ('McMaster-Carr','https://www.mcmaster.com'),
    'mcmaster': ('McMaster-Carr','https://www.mcmaster.com'),
    'pasternack': ('Pasternack','https://www.pasternack.com'),
    'rsonline': ('RS Online','https://www.rs-online.com'),
    'skygeek': ('SkyGeek','https://www.skygeek.com'),
    'swagelok': ('Swagelok','https://www.swagelok.com'),
    'uline': ('Uline','https://www.uline.com'),
}
def norm_company(raw):
    if not raw: return ('', None)
    s = str(raw).strip()
    key = re.sub(r'[^a-z]', '', s.lower())
    if key in VENDOR_URLS:
        return VENDOR_URLS[key]
    # leave display name as-is, no landing page for unknown / "Not On List"
    return (s, None)

# ---------- Type normalization ----------
def norm_type(raw):
    if not raw: return 'Hardware/Tools'
    s=str(raw).strip().lower()
    if 'consumable' in s: return 'Consumable'
    if 'safety' in s: return 'Safety'
    if 'program' in s: return 'Program'
    if 'tool' in s and 'hardware' not in s: return 'Tool'
    if 'hardware' in s: return 'Hardware/Tools'
    return str(raw).strip().title()

def is_url(v):
    return isinstance(v,str) and v.strip().lower().startswith(('http://','https://'))

def best_link(rec):
    if rec.get('LinkTarget') and is_url(rec['LinkTarget']): return rec['LinkTarget'].strip()
    if is_url(rec.get('Hyperlink')): return rec['Hyperlink'].strip()
    if rec.get('ItemLink') and is_url(rec['ItemLink']): return rec['ItemLink'].strip()
    return ''

# attach parsed fields
for rec in rows:
    fn,ln,d,t = parse_requestor(rec['Requestor'])
    rec['_fn']=fn; rec['_ln']=ln; rec['_date']=d; rec['_time']=t

# ---------- Build user list ----------
people = OrderedDict()  # (fn,ln) -> info
for rec in rows:
    people.setdefault((rec['_fn'],rec['_ln']), 0)
    people[(rec['_fn'],rec['_ln'])] += 1

ESS_PEOPLE = {('Mike','Greere'), ('Gerald','Jackson')}
TITLES = {
    ('Mike','Greere'):'ESS Lead', ('Gerald','Jackson'):'Procurement Admin',
    ('Morgan','Elliott'):'MFG Engineer', ('Tom','DeMars'):'Integration Tech',
    ('Sydney','Thompson'):'MFG Engineer', ('Roman','Davtyan'):'Integration Tech',
    ('Sven','Alagic'):'Test Engineer', ('Darrin','Loehr'):'Facilities Lead',
    ('Maureen','Kapler'):'Quality Engineer', ('Mike','Zavertaniy'):'Wiring Tech',
    ('Jeremy','Winter'):'Technician', ('Dan','Past'):'Technician',
    ('Todd','Leadens'):'Technician', ('Greg','Zdeb'):'Engineer',
    ('Chufu','Lin'):'Engineer', ('Macleod','Macleod'):'Technician', ('Hanson','Hanson'):'Technician',
}

def title_for(p): return TITLES.get(p, 'Engineer')

# usernames unique
usernames = {}
used = set(['admin'])
for (fn,ln) in people:
    base = re.sub(r'[^a-z0-9]','', fn.lower()) or re.sub(r'[^a-z0-9]','', ln.lower())
    uname = base
    if uname in used:
        uname = base + re.sub(r'[^a-z0-9]','', ln.lower())[:1]
    i=2
    while uname in used:
        uname = base+str(i); i+=1
    used.add(uname); usernames[(fn,ln)] = uname

def default_pwd(fn,ln):
    if not fn or not ln: return 'password'
    return f"{ln.strip()}{fn.strip()[0]}".replace(' ','')

users = []
for idx,(p,cnt) in enumerate(people.items(), start=1):
    fn,ln = p
    role = 'ESS' if p in ESS_PEOPLE else 'Employee'
    users.append({
        'id': f'h-u{idx}',
        'firstName': fn, 'lastName': ln, 'jobTitle': title_for(p),
        'role': role, 'username': usernames[p],
        'pwd': default_pwd(fn,ln),
    })

def display_name(p):
    fn,ln = p
    return f"{fn} {ln} - {title_for(p)}"

# ---------- Group rows into PRs by raw requestor string ----------
groups = OrderedDict()
for rec in rows:
    groups.setdefault(str(rec['Requestor']), []).append(rec)

def iso(d, t=None):
    if not d: d='2025-01-01'
    t = t or '09:00:00'
    return f"{d}T{t}"

requests = []
ridx = 1
for raw, recs in groups.items():
    p = (recs[0]['_fn'], recs[0]['_ln'])
    d = recs[0]['_date']; t = recs[0]['_time']
    created = iso(d,t)
    created_dt = datetime.fromisoformat(created)
    needed = (created_dt + timedelta(days=14)).isoformat()
    # build cell -> projectCode
    cells = [str(r['Buildcell']).strip().title() for r in recs if r['Buildcell']]
    project = Counter(cells).most_common(1)[0][0] if cells else 'Historic Import'
    items = []
    ordered_count = 0
    for i, rec in enumerate(recs, start=1):
        vname, vurl = norm_company(rec['Company'])
        cost = rec['Cost'] if isinstance(rec['Cost'],(int,float)) else 0.0
        qty = rec['Qty'] if isinstance(rec['Qty'],(int,float)) else 1
        try: qty = float(qty)
        except: qty = 1.0
        unit = str(rec['Unit']).strip() if rec['Unit'] else 'ea'
        model = rec['Model']
        model = str(int(model)) if isinstance(model,float) and model.is_integer() else (str(model).strip() if model is not None else '')
        ordered = str(rec['Ordered']).strip().lower()=='yes'
        received = str(rec['Received']).strip().lower()=='yes'
        if ordered: ordered_count += 1
        items.append({
            'id': f'REQ-H{ridx:04d}-i{i}',
            'name': str(rec['Item']).strip() if rec['Item'] else 'Item',
            'description': str(rec['Details']).strip() if rec['Details'] else '',
            'vendor': vname,
            'mfgPartNumber': model,
            'url': best_link(rec),
            'quantity': qty,
            'unitType': unit,
            'pricePerUnit': round(float(cost),2),
            'itemType': norm_type(rec['Type']),
            'buildCell': str(rec['Buildcell']).strip().title() if rec['Buildcell'] else '',
            'vendorUrl': vurl or '',
            'ordered': ordered,
            'received': received,
        })
    total = round(sum(it['quantity']*it['pricePerUnit'] for it in items),2)
    if ordered_count==0: status='PENDING'
    elif ordered_count==len(items): status='ORDERED'
    else: status='ORDERED'
    timeline = [{
        'id': f'REQ-H{ridx:04d}-e1', 'role':'Employee', 'actorName': display_name(p),
        'action':'Submitted', 'timestamp': created
    }]
    if status=='ORDERED':
        ts = (created_dt + timedelta(days=1)).isoformat()
        timeline.append({'id': f'REQ-H{ridx:04d}-e2', 'role':'ESS',
            'actorName':'Mike Greere - ESS Lead', 'action':'Ordered', 'timestamp': ts,
            'note':'Imported from Master PR Record'})
    requests.append({
        'id': f'REQ-H{ridx:04d}', 'projectCode': project, 'requesterName': display_name(p),
        'createdAt': created, 'updatedAt': created, 'neededByDate': needed,
        'priority':'NORMAL', 'status': status, 'items': items,
        'messages': [], 'approvalTimeline': timeline, 'totalAmount': total,
        'notes': f'Historic import (orig: {raw})', 'submittedDate': d or '', 'submittedTime': t or '',
    })
    ridx += 1

# ---------- Emit TypeScript ----------
def js(v):
    return json.dumps(v, ensure_ascii=False)

lines = []
lines.append("// AUTO-GENERATED from historicData/Submitted PRs/MasterList/Master_PR_Record.xlsm")
lines.append("// Do not edit by hand. Regenerate via scripts/generate_seed.py")
lines.append("import { User, PurchaseRequest, RequestStatus, Priority } from '../types';")
lines.append("")
lines.append("export type SeedUser = Omit<User, 'passwordHash'> & { defaultPassword: string };")
lines.append("")
lines.append("export const HISTORIC_USERS: SeedUser[] = [")
for u in users:
    lines.append("  { id: %s, firstName: %s, lastName: %s, jobTitle: %s, role: %s, username: %s, isDefaultPassword: true, defaultPassword: %s }," % (
        js(u['id']), js(u['firstName']), js(u['lastName']), js(u['jobTitle']), js(u['role']), js(u['username']), js(u['pwd'])))
lines.append("];")
lines.append("")

STAT = {'PENDING':'RequestStatus.PENDING','ORDERED':'RequestStatus.ORDERED','RECEIVED':'RequestStatus.RECEIVED'}
lines.append("export const HISTORIC_REQUESTS: PurchaseRequest[] = [")
for r in requests:
    lines.append("  {")
    lines.append("    id: %s," % js(r['id']))
    lines.append("    projectCode: %s," % js(r['projectCode']))
    lines.append("    requesterName: %s," % js(r['requesterName']))
    lines.append("    createdAt: %s," % js(r['createdAt']))
    lines.append("    updatedAt: %s," % js(r['updatedAt']))
    lines.append("    neededByDate: %s," % js(r['neededByDate']))
    lines.append("    submittedDate: %s," % js(r['submittedDate']))
    lines.append("    submittedTime: %s," % js(r['submittedTime']))
    lines.append("    priority: Priority.NORMAL,")
    lines.append("    status: %s," % STAT[r['status']])
    lines.append("    totalAmount: %s," % json.dumps(r['totalAmount']))
    lines.append("    notes: %s," % js(r['notes']))
    lines.append("    messages: [],")
    lines.append("    items: [")
    for it in r['items']:
        lines.append("      { id: %s, name: %s, description: %s, vendor: %s, mfgPartNumber: %s, url: %s, quantity: %s, unitType: %s, pricePerUnit: %s, itemType: %s, buildCell: %s, vendorUrl: %s, ordered: %s, received: %s }," % (
            js(it['id']), js(it['name']), js(it['description']), js(it['vendor']), js(it['mfgPartNumber']),
            js(it['url']), json.dumps(it['quantity']), js(it['unitType']), json.dumps(it['pricePerUnit']),
            js(it['itemType']), js(it['buildCell']), js(it['vendorUrl']),
            'true' if it['ordered'] else 'false', 'true' if it['received'] else 'false'))
    lines.append("    ],")
    lines.append("    approvalTimeline: [")
    for e in r['approvalTimeline']:
        note = (", note: %s" % js(e['note'])) if e.get('note') else ""
        lines.append("      { id: %s, role: %s, actorName: %s, action: %s, timestamp: %s%s }," % (
            js(e['id']), js(e['role']), js(e['actorName']), js(e['action']), js(e['timestamp']), note))
    lines.append("    ],")
    lines.append("  },")
lines.append("];")
lines.append("")

out = "\n".join(lines)
with open(r'services\historicData.ts','w',encoding='utf-8') as f:
    f.write(out)

print(f"Users: {len(users)}  Requests: {len(requests)}  Items: {sum(len(r['items']) for r in requests)}")
print("Vendors with landing pages used:", sorted({it['vendorUrl'] for r in requests for it in r['items'] if it['vendorUrl']}))
print("Wrote services/historicData.ts")
